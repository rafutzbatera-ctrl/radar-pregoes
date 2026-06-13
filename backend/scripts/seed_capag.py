"""Popula as tabelas CAPAG (Tesouro Nacional / SICONFI) — rodar manual/mensal.

Baixa, em uma req/s, com User-Agent identificável:
  1. /entes do SICONFI (JSON paginado, 5000/pág) → casa CNPJ→cod_ibge.
  2. XLSX anual da CAPAG de MUNICÍPIOS e de ESTADOS, resolvendo o resource mais
     novo via CKAN package_show (datasets capag-municipios e capag-estados).

Parseia o XLSX com openpyxl (read_only=True, ~22MB), aba "Prévia da CAPAG", e
popula capag_entes/capag_notas com INSERT OR REPLACE (idempotente). Tolerante a
coluna ausente/linha suja (try/continua). Abre o mesmo data/radar.db via app.db.

NÃO baixa nada no request do app — só este script popula. Uso:
    python -m scripts.seed_capag            (de dentro de backend/)
    python scripts/seed_capag.py
"""
import io
import sys
import time
from pathlib import Path

import httpx
from openpyxl import load_workbook

# permitir rodar tanto como `python -m scripts.seed_capag` quanto direto
sys.path.insert(0, str(Path(__file__).resolve().parents[1]))

from app import db, settings  # noqa: E402

UA = settings.USER_AGENT  # "RadarPregoes/0.1 (uso pessoal)"
INTERVALO = 1.0  # ~1 req/s (gentileza com API pública — CLAUDE.md §2.6)
TIMEOUT = 180.0  # XLSX de ~22MB

CKAN = "https://www.tesourotransparente.gov.br/ckan/api/3/action/package_show"
ENTES = "https://apidatalake.tesouro.gov.br/ords/siconfi/tt/entes"
ABA = "Prévia da CAPAG"

# mapeamento (nome normalizado da coluna do XLSX) → chave interna. Vários
# cabeçalhos possíveis por tolerância a variações entre anos da planilha.
COLS = {
    "codigo municipio completo": "cod_ibge",
    "cod. ibge": "cod_ibge",
    "cod ibge": "cod_ibge",
    "codigo ibge": "cod_ibge",
    "nome_municipio": "municipio",
    "nome municipio": "municipio",
    "municipio": "municipio",
    "ente": "municipio",
    "uf": "uf",
    "capag": "nota",
    "indicador 1": "ind1",
    "nota 1": "nota1",
    "indicador 2": "ind2",
    "nota 2": "nota2",
    "indicador 3": "ind3",
    "nota 3": "nota3",
    "icf": "icf",
    "origem da nota final": "origem",
    "origem da nota": "origem",
}


def _log(*a):
    print(*a, flush=True)


def _normalizar_cabecalho(s) -> str:
    return str(s or "").strip().lower()


def _http() -> httpx.Client:
    return httpx.Client(headers={"User-Agent": UA}, timeout=TIMEOUT,
                        follow_redirects=True)


def _esperar():
    time.sleep(INTERVALO)


def resolver_xlsx(cli: httpx.Client, dataset_id: str) -> str | None:
    """CKAN package_show → URL do resource .xlsx mais recente."""
    _log(f"  · resolvendo XLSX de '{dataset_id}' via CKAN…")
    r = cli.get(CKAN, params={"id": dataset_id})
    r.raise_for_status()
    res = (r.json().get("result") or {}).get("resources") or []
    xlsx = [x for x in res if str(x.get("url", "")).lower().endswith(".xlsx")]
    if not xlsx:
        _log(f"    ! nenhum .xlsx em '{dataset_id}'")
        return None

    def chave(x):
        # ordena pela data mais nova disponível no metadado
        return (x.get("last_modified") or x.get("created")
                or x.get("metadata_modified") or "")

    escolhido = max(xlsx, key=chave)
    _log(f"    → {escolhido.get('url')}")
    return escolhido.get("url")


def baixar_entes(con, cli: httpx.Client) -> int:
    """Baixa todos os /entes (5000/pág) e popula capag_entes (cnpj→cod_ibge)."""
    _log("entes (SICONFI): baixando…")
    offset = 0
    total = 0
    while True:
        _esperar()
        r = cli.get(ENTES, params={"offset": offset})
        r.raise_for_status()
        corpo = r.json()
        itens = corpo.get("items") or []
        for it in itens:
            try:
                cnpj = "".join(ch for ch in str(it.get("cnpj") or "") if ch.isdigit())
                cod = str(it.get("cod_ibge") or "").strip()
                if not cnpj or not cod:
                    continue
                con.execute(
                    "INSERT OR REPLACE INTO capag_entes"
                    "(cnpj, cod_ibge, ente, uf, esfera) VALUES (?,?,?,?,?)",
                    (cnpj, cod, it.get("ente"), it.get("uf"), it.get("esfera")),
                )
                total += 1
            except Exception as exc:  # noqa: BLE001 — linha suja não derruba o lote
                _log(f"    ! ente ignorado: {exc}")
                continue
        con.commit()
        _log(f"  · {total} entes acumulados (offset={offset})")
        if not corpo.get("hasMore"):
            break
        offset += len(itens) if itens else 5000
    _log(f"entes: {total} gravados.")
    return total


def _to_float(v):
    if v is None or v == "":
        return None
    try:
        return float(str(v).replace(".", "").replace(",", ".")) \
            if isinstance(v, str) and "," in v else float(v)
    except (TypeError, ValueError):
        return None


def parsear_xlsx(con, conteudo: bytes, esfera: str) -> int:
    """Parseia a aba 'Prévia da CAPAG' e popula capag_notas."""
    wb = load_workbook(io.BytesIO(conteudo), read_only=True, data_only=True)
    aba = ABA if ABA in wb.sheetnames else wb.sheetnames[0]
    if ABA not in wb.sheetnames:
        _log(f"    ! aba '{ABA}' ausente; usando '{aba}'")
    ws = wb[aba]

    linhas = ws.iter_rows(values_only=True)
    try:
        cabecalho = next(linhas)
    except StopIteration:
        return 0
    # índice de cada coluna conhecida
    idx = {}
    for i, nome in enumerate(cabecalho):
        chave = COLS.get(_normalizar_cabecalho(nome))
        if chave and chave not in idx:
            idx[chave] = i
    if "cod_ibge" not in idx or "nota" not in idx:
        _log(f"    ! cabeçalho sem cod_ibge/CAPAG (cols: {list(idx)}) — pulando")
        wb.close()
        return 0

    def pega(linha, chave):
        i = idx.get(chave)
        if i is None or i >= len(linha):
            return None
        return linha[i]

    total = 0
    for linha in linhas:
        try:
            cod = pega(linha, "cod_ibge")
            cod = str(int(cod)) if isinstance(cod, float) else str(cod or "").strip()
            nota = pega(linha, "nota")
            if not cod or not nota:
                continue
            con.execute(
                "INSERT OR REPLACE INTO capag_notas"
                "(cod_ibge, municipio, uf, nota, ind1, nota1, ind2, nota2,"
                " ind3, nota3, icf, origem, esfera)"
                " VALUES (?,?,?,?,?,?,?,?,?,?,?,?,?)",
                (
                    cod,
                    pega(linha, "municipio"),
                    pega(linha, "uf"),
                    str(nota).strip(),
                    _to_float(pega(linha, "ind1")), pega(linha, "nota1"),
                    _to_float(pega(linha, "ind2")), pega(linha, "nota2"),
                    _to_float(pega(linha, "ind3")), pega(linha, "nota3"),
                    pega(linha, "icf"),
                    pega(linha, "origem"),
                    esfera,
                ),
            )
            total += 1
        except Exception as exc:  # noqa: BLE001 — linha suja não derruba o lote
            _log(f"    ! linha ignorada: {exc}")
            continue
    con.commit()
    wb.close()
    return total


def baixar_notas(con, cli: httpx.Client, dataset_id: str, esfera: str) -> int:
    url = resolver_xlsx(cli, dataset_id)
    if not url:
        return 0
    _log(f"  · baixando XLSX ({dataset_id})… (pode levar um tempo, ~22MB)")
    _esperar()
    r = cli.get(url)
    r.raise_for_status()
    _log(f"  · {len(r.content) // 1024} KB baixados; parseando…")
    n = parsear_xlsx(con, r.content, esfera)
    _log(f"  · {n} notas gravadas ({dataset_id}).")
    return n


def main():
    _log("=== seed CAPAG (Tesouro Nacional / SICONFI) ===")
    con = db.abrir()
    try:
        with _http() as cli:
            baixar_entes(con, cli)
            _log("notas CAPAG — municípios:")
            baixar_notas(con, cli, "capag-municipios", "M")
            _log("notas CAPAG — estados:")
            baixar_notas(con, cli, "capag-estados", "E")
        n_entes = con.execute("SELECT COUNT(*) FROM capag_entes").fetchone()[0]
        n_notas = con.execute("SELECT COUNT(*) FROM capag_notas").fetchone()[0]
        _log(f"=== concluído: {n_entes} entes, {n_notas} notas ===")
    finally:
        con.close()


if __name__ == "__main__":
    main()
